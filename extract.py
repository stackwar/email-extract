import imaplib
import email
import base64
import hashlib
import hmac
import io
import json
import re
import sys
import time
import tkinter as tk
from datetime import datetime
from email.header import decode_header
from pathlib import Path
from tkinter import messagebox

import easyocr
from openpyxl import Workbook, load_workbook

ocr_reader = None
LICENSE_SECRET = b"rider_extract_2026_sk_x9f3m"
LICENSE_FILE = "license.key"
LAST_RUN_FILE = ".last_run"


def verify_license(code):
    try:
        raw = base64.b64decode(code.strip())
        if len(raw) != 18:
            return False, "授权码格式无效"
        expire_date = raw[:10].decode("utf-8")
        sig = raw[10:]
        expected = hmac.HMAC(LICENSE_SECRET, raw[:10], hashlib.sha256).digest()[:8]
        if not hmac.compare_digest(sig, expected):
            return False, "授权码无效"
        expire = datetime.strptime(expire_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
        if datetime.now() > expire:
            return False, f"授权码已过期 ({expire_date})"
        return True, expire_date
    except Exception:
        return False, "授权码格式错误"


def check_time_rollback():
    path = Path(LAST_RUN_FILE)
    today = datetime.now().strftime("%Y-%m-%d")
    if path.exists():
        last_date = path.read_text().strip()
        if last_date > today:
            messagebox.showerror("错误", "检测到系统时间异常，请恢复正确的系统时间后重试")
            sys.exit(1)
    path.write_text(today)


def show_license_dialog(error_msg=None):
    result = {"code": None}

    dialog = tk.Tk()
    dialog.title("授权验证")
    dialog.resizable(False, False)
    dialog.attributes("-topmost", True)

    window_width, window_height = 380, 180
    screen_width = dialog.winfo_screenwidth()
    screen_height = dialog.winfo_screenheight()
    x = (screen_width - window_width) // 2
    y = (screen_height - window_height) // 2
    dialog.geometry(f"{window_width}x{window_height}+{x}+{y}")

    if error_msg:
        tk.Label(dialog, text=error_msg, fg="red").pack(pady=(10, 0))

    tk.Label(dialog, text="请输入授权码：").pack(pady=(10, 5))
    entry = tk.Entry(dialog, width=40)
    entry.pack(pady=5)
    entry.focus_set()

    def on_submit(event=None):
        result["code"] = entry.get().strip()
        dialog.destroy()

    def on_cancel():
        dialog.destroy()

    entry.bind("<Return>", on_submit)
    btn_frame = tk.Frame(dialog)
    btn_frame.pack(pady=15)
    tk.Button(btn_frame, text="验证", width=10, command=on_submit).pack(side=tk.LEFT, padx=10)
    tk.Button(btn_frame, text="退出", width=10, command=on_cancel).pack(side=tk.LEFT, padx=10)

    dialog.protocol("WM_DELETE_WINDOW", on_cancel)
    dialog.mainloop()

    return result["code"]


def check_license():
    check_time_rollback()

    path = Path(LICENSE_FILE)
    if path.exists():
        code = path.read_text().strip()
        valid, msg = verify_license(code)
        if valid:
            return
        path.unlink()
        error_msg = f"授权码已失效: {msg}"
    else:
        error_msg = None

    while True:
        code = show_license_dialog(error_msg)
        if not code:
            sys.exit(0)
        valid, msg = verify_license(code)
        if valid:
            path.write_text(code)
            messagebox.showinfo("授权验证", f"验证通过，有效期至 {msg}")
            return
        error_msg = f"错误: {msg}"


def get_ocr_reader():
    global ocr_reader
    if ocr_reader is None:
        ocr_reader = easyocr.Reader(['ch_sim', 'en'], gpu=False)
    return ocr_reader

CONFIG_FILE = "config.json"


def load_config():
    path = Path(CONFIG_FILE)
    if not path.exists():
        print(f"错误: 未找到配置文件 {CONFIG_FILE}，请参考 config.example.json 创建")
        sys.exit(1)
    config = json.loads(path.read_text())
    if not config.get("email_addr") or not config.get("auth_code"):
        print(f"错误: {CONFIG_FILE} 中缺少 email_addr 或 auth_code")
        sys.exit(1)
    return config


config = load_config()
IMAP_SERVER = config.get("imap_server", "imap.qq.com")
EMAIL_ADDR = config["email_addr"]
AUTH_CODE = config["auth_code"]
SUBJECT_KEYWORD = "聊天记录"
OUTPUT_FILE = "骑手信息.xlsx"
PROCESSED_FILE = "processed_ids.json"


def load_processed_ids():
    path = Path(PROCESSED_FILE)
    if path.exists():
        return set(json.loads(path.read_text()))
    return set()


def save_processed_ids(ids):
    Path(PROCESSED_FILE).write_text(json.dumps(sorted(ids)))


def decode_str(s):
    if s is None:
        return ""
    parts = decode_header(s)
    result = ""
    for content, charset in parts:
        if isinstance(content, bytes):
            result += content.decode(charset or "utf-8", errors="ignore")
        else:
            result += content
    return result


def get_email_body(msg):
    body = ""
    if msg.is_multipart():
        for part in msg.walk():
            if part.get_content_type() == "text/plain":
                payload = part.get_payload(decode=True)
                charset = part.get_content_charset() or "utf-8"
                body += payload.decode(charset, errors="ignore")
    else:
        payload = msg.get_payload(decode=True)
        charset = msg.get_content_charset() or "utf-8"
        body = payload.decode(charset, errors="ignore")
    return body


def get_inline_images(msg):
    images = []
    if not msg.is_multipart():
        return images
    for part in msg.walk():
        content_type = part.get_content_type()
        if content_type.startswith("image/"):
            payload = part.get_payload(decode=True)
            if payload:
                images.append(payload)
    return images


def ocr_images_to_text(images):
    reader = get_ocr_reader()
    all_text = []
    for img_data in images:
        try:
            results = reader.readtext(img_data)
            lines = [r[1] for r in results]
            all_text.append("\n".join(lines))
        except Exception:
            continue
    return "\n".join(all_text)


def extract_date_from_text(text):
    date_pattern = re.compile(r'(\d{4})[年/\-.](\d{1,2})[月/\-.](\d{1,2})')
    match = date_pattern.search(text)
    if match:
        y, m, d = match.group(1), match.group(2).zfill(2), match.group(3).zfill(2)
        return f"{y}-{m}-{d}"
    return None


def parse_records(text, fallback_date="未知日期"):
    records = []
    current_date = fallback_date
    date_line_pattern = re.compile(r'—+\s*(\d{4}[-/\.]\d{1,2}[-/\.]\d{1,2})\s*—+')
    date_cn_pattern = re.compile(r'(\d{4})[年](\d{1,2})[月](\d{1,2})[日]')

    sections = re.split(r'(—+\s*\d{4}[-/\.]\d{1,2}[-/\.]\d{1,2}\s*—+)', text)

    for section in sections:
        date_match = date_line_pattern.search(section)
        if date_match:
            raw = date_match.group(1)
            parts = re.split(r'[-/\.]', raw)
            current_date = f"{parts[0]}-{parts[1].zfill(2)}-{parts[2].zfill(2)}"
            continue

        blocks = re.split(r'(?=(?:骑手)?姓名[:：\s])', section)
        for block in blocks:
            if "姓名" not in block:
                continue
            try:
                cn_date_match = date_cn_pattern.search(block)
                if cn_date_match:
                    current_date = f"{cn_date_match.group(1)}-{cn_date_match.group(2).zfill(2)}-{cn_date_match.group(3).zfill(2)}"

                record = {"日期": current_date}
                name_match = re.search(r'(?:骑手)?姓名[:：\s]*(.+)', block)
                phone_match = re.search(r'(?:骑手)?电话[:：\s]*(.+)', block)
                site_match = re.search(r'入职站点[:：\s]*(.+)', block)
                housing_match = re.search(r'是否\s*住宿[:：\s]*(.+)', block)
                job_match = re.search(r'兼职[/／1]全职[:：\s]*(.+)', block)
                remark_match = re.search(r'备注[:：\s]*(.+)', block)

                record["骑手姓名"] = name_match.group(1).strip() if name_match else ""
                record["骑手电话"] = phone_match.group(1).strip() if phone_match else ""
                record["入职站点"] = site_match.group(1).strip() if site_match else ""
                record["是否住宿"] = housing_match.group(1).strip() if housing_match else ""
                record["兼职/全职"] = job_match.group(1).strip() if job_match else ""
                record["备注"] = remark_match.group(1).strip() if remark_match else ""

                if record["骑手姓名"]:
                    records.append(record)
            except Exception as e:
                print(f"警告: 解析记录时出错，已跳过。错误: {e}")
    return records


HEADERS = ["日期", "骑手姓名", "骑手电话", "入职站点", "是否住宿", "兼职/全职", "备注"]


def get_or_create_sheet(wb, sheet_name):
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)
    ws.append(HEADERS)
    return ws


def append_to_excel(records):
    path = Path(OUTPUT_FILE)
    if path.exists():
        wb = load_workbook(OUTPUT_FILE)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    existing_keys = set()
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and len(row) >= 3:
                existing_keys.add((str(row[1]).strip(), str(row[2]).strip()))

    new_count = 0
    for r in records:
        key = (r["骑手姓名"], r["骑手电话"])
        if key in existing_keys:
            print(f"  重复跳过: {r['骑手姓名']} ({r['骑手电话']})")
            continue

        date_str = r["日期"]
        if re.match(r'\d{4}-\d{2}', date_str):
            sheet_name = date_str[:7]
        else:
            sheet_name = "未知日期"

        ws = get_or_create_sheet(wb, sheet_name)
        ws.append([r["日期"], r["骑手姓名"], r["骑手电话"], r["入职站点"], r["是否住宿"], r["兼职/全职"], r["备注"]])
        existing_keys.add(key)
        new_count += 1

    wb.save(OUTPUT_FILE)
    return new_count


def connect_and_login():
    try:
        mail = imaplib.IMAP4_SSL(IMAP_SERVER)
    except Exception as e:
        print(f"错误: 无法连接到 {IMAP_SERVER}。{e}")
        return None

    try:
        mail.login(EMAIL_ADDR, AUTH_CODE)
    except imaplib.IMAP4.error as e:
        print(f"错误: 登录失败。请检查邮箱地址和授权码。{e}")
        return None

    mail.select("INBOX")
    return mail


def process_new_emails(mail):
    _, data = mail.uid('search', 'UTF-8', 'SUBJECT', SUBJECT_KEYWORD.encode('utf-8'))
    mail_uids = data[0].split()

    if not mail_uids:
        return 0

    processed_ids = load_processed_ids()
    total_records = 0

    for uid_bytes in reversed(mail_uids):
        uid = uid_bytes.decode()
        if uid in processed_ids:
            continue

        try:
            _, msg_data = mail.uid('fetch', uid_bytes, "(RFC822)")
            msg = email.message_from_bytes(msg_data[0][1])
        except Exception as e:
            print(f"警告: 获取邮件 {uid} 失败，已跳过。{e}")
            continue

        subject = decode_str(msg["Subject"])
        if SUBJECT_KEYWORD not in subject:
            processed_ids.add(uid)
            continue

        try:
            date_str = email.utils.parsedate_to_datetime(msg["Date"]).strftime("%Y-%m-%d")
        except Exception:
            date_str = "未知日期"

        body = get_email_body(msg)
        images = get_inline_images(msg)
        ocr_text = ""

        if images:
            ocr_text = ocr_images_to_text(images)

        records = parse_records(body, fallback_date=date_str)
        if not records and ocr_text:
            records = parse_records(ocr_text, fallback_date=date_str)

        if records:
            added = append_to_excel(records)
            total_records += added
            if added < len(records):
                print(f"邮件 {uid} ({date_str}): 提取 {len(records)} 条记录，实际写入 {added} 条（{len(records) - added} 条重复）")
            else:
                print(f"邮件 {uid} ({date_str}): 提取 {len(records)} 条记录")
        else:
            print(f"邮件 {uid} ({date_str}): 未提取到骑手信息")

        processed_ids.add(uid)

    save_processed_ids(processed_ids)
    return total_records


def main():
    mail = connect_and_login()
    if not mail:
        sys.exit(1)

    total = process_new_emails(mail)
    mail.logout()

    if total > 0:
        print(f"完成。共提取 {total} 条新记录，已保存到 {OUTPUT_FILE}")
    else:
        print("没有新的记录需要处理")


def watch():
    print(f"[{datetime.now().strftime('%H:%M:%S')}] 开始监听邮箱新邮件...")
    print("按 Ctrl+C 停止\n")

    while True:
        mail = connect_and_login()
        if not mail:
            print("连接失败，30秒后重试...")
            time.sleep(30)
            continue

        try:
            total = process_new_emails(mail)
            if total > 0:
                print(f"[{datetime.now().strftime('%H:%M:%S')}] 新增 {total} 条记录到 {OUTPUT_FILE}")

            # IMAP IDLE - 等待新邮件通知
            tag = mail._new_tag()
            mail.send(tag + b' IDLE\r\n')
            mail.readline()  # + idling

            print(f"[{datetime.now().strftime('%H:%M:%S')}] 等待新邮件...")

            while True:
                try:
                    mail.sock.settimeout(1680)  # 28分钟超时，避免服务器断开(RFC建议29分钟内刷新)
                    line = mail.readline()
                    if b'EXISTS' in line:
                        print(f"[{datetime.now().strftime('%H:%M:%S')}] 收到新邮件通知")
                        break
                    if b'BYE' in line:
                        break
                except (TimeoutError, OSError):
                    break

            # 退出 IDLE
            try:
                mail.send(b'DONE\r\n')
                mail.readline()
            except Exception:
                pass

        except KeyboardInterrupt:
            print("\n停止监听")
            try:
                mail.logout()
            except Exception:
                pass
            return
        except Exception as e:
            print(f"[{datetime.now().strftime('%H:%M:%S')}] 连接异常: {e}，10秒后重连...")
            time.sleep(10)


if __name__ == "__main__":
    check_license()
    if len(sys.argv) > 1 and sys.argv[1] == "watch":
        watch()
    else:
        main()
